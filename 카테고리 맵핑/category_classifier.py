import csv
import os
import openpyxl

# 카테고리2 고유값을 대분류로 매핑
CATEGORY_MAP = {
    # Outer
    '가디건': 'Outer', '코트': 'Outer', '바람막이': 'Outer', '패딩': 'Outer', '케이프': 'Outer',
    '더플코트': 'Outer', '트렌치코트': 'Outer', '집업재킷': 'Outer', '재킷': 'Outer', '트럭커': 'Outer',
    '블루종': 'Outer', '야상': 'Outer', '발마칸재킷': 'Outer', '모피코트': 'Outer', '항공점퍼': 'Outer',
    '점퍼': 'Outer',
    # Top
    '티셔츠': 'Top', '니트': 'Top', '후드': 'Top', '블라우스': 'Top', '셔츠': 'Top', '맨투맨': 'Top',
    '반팔티': 'Top', '베스트': 'Top', '조끼': 'Top', '스웨터': 'Top', '폴로셔츠': 'Top',
    '파자마상의': 'Top', '터틀넥': 'Top', '스웨트셔츠': 'Top', '가디건': 'Top',
    # Bottom
    '바지': 'Bottom', '슬랙스': 'Bottom', '청바지': 'Bottom', '스커트': 'Bottom', '배기팬츠': 'Bottom',
    '와이드팬츠': 'Bottom', '치노팬츠': 'Bottom', '플레어스커트': 'Bottom', '플리츠스커트': 'Bottom',
    '트럼펫스커트': 'Bottom', '티어드스커트': 'Bottom', '러플스커트': 'Bottom', '잠옷바지': 'Bottom',
    '청반바지': 'Bottom', '스웨트팬츠': 'Bottom',
    # Dress&Setup
    '원피스': 'Dress&Setup', '드레스': 'Dress&Setup', '셔츠원피스': 'Dress&Setup', '튜닉원피스': 'Dress&Setup',
    '슬립드레스': 'Dress&Setup', '멜빵원피스': 'Dress&Setup', '피케원피스': 'Dress&Setup',
    '파자마드레스': 'Dress&Setup', '점프수트': 'Dress&Setup', '멜빵바지': 'Dress&Setup',
    '니트원피스': 'Dress&Setup', '페플럼드레스': 'Dress&Setup', '뷔스티에': 'Dress&Setup', '데님원피스': 'Dress&Setup',
    # Accessory
    '모자': 'Accessory', '비니': 'Accessory', '귀마개': 'Accessory', '타이츠스타킹': 'Accessory',
    '샌들': 'Accessory', '스니커즈': 'Accessory', '슬립온': 'Accessory', '어그부츠': 'Accessory',
    # ETC (기타)
}

# 키워드 기반 분류 보조 (매핑에 없는 경우)
KEYWORDS = [
    ('Outer', ['코트', '패딩', '재킷', '점퍼', '블루종', '야상', '트렌치', '케이프', '모피', '트럭커', '발마칸', '항공']),
    ('Top', ['티셔츠', '니트', '후드', '블라우스', '셔츠', '맨투맨', '베스트', '조끼', '스웨터', '폴로', '파자마상', '터틀넥']),
    ('Bottom', ['바지', '슬랙스', '청바지', '스커트', '팬츠', '배기', '와이드', '치노', '플레어', '플리츠', '트럼펫', '티어드', '러플', '잠옷', '청반']),
    ('Dress&Setup', ['원피스', '드레스', '점프수트', '멜빵', '튜닉', '슬립', '피케', '파자마', '뷔스티에', '데님', '페플럼']),
    ('Accessory', ['모자', '비니', '귀마개', '타이츠', '샌들', '스니커즈', '슬립온', '어그', '가방', '백팩', '장갑', '머플러', '스카프']),
]

def classify(category2: str, category1: str = "") -> str:
    c1 = (category1 or "").strip()
    c2 = (category2 or "").strip()
    # 1. 카테고리1이 잡화, 팔찌, 귀걸이, 목걸이, 브로치면 Accessory
    if c1 in ['잡화', '팔찌', '귀걸이', '목걸이', '브로치']:
        return 'Accessory'
    # 2. 카테고리1이 탑이면 Top
    if c1 == '탑':
        return 'Top'
    if c2 in CATEGORY_MAP:
        return CATEGORY_MAP[c2]
    for big, keywords in KEYWORDS:
        for kw in keywords:
            if kw in c2:
                return big
    if c1:
        for big, keywords in KEYWORDS:
            for kw in keywords:
                if kw in c1:
                    return big
    return 'ETC'

input_path = os.path.join(os.path.dirname(__file__), '카테고리 1, 2.csv')
output_path = os.path.join(os.path.dirname(__file__), '카테고리_분류결과.csv')

def main():
    with open(input_path, newline='', encoding='utf-8') as infile, \
         open(output_path, 'w', newline='', encoding='utf-8') as outfile:
        reader = csv.reader(infile)
        writer = csv.writer(outfile)
        header = next(reader)
        writer.writerow(header + ['대분류'])
        for row in reader:
            if len(row) < 2:
                continue
            cat1, cat2 = row[0], row[1]
            big = classify(cat2, cat1)
            writer.writerow(row + [big])

def save_unique_cp949(input_path, output_path):
    seen = set()
    rows = []
    with open(input_path, newline='', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        header = next(reader)
        for row in reader:
            if len(row) < 3:
                continue
            key = tuple(row[:3])
            if key not in seen:
                seen.add(key)
                rows.append(row[:3])
    with open(output_path, 'w', newline='', encoding='cp949') as outfile:
        writer = csv.writer(outfile)
        writer.writerow(['카테고리1', '카테고리2', '대분류'])
        writer.writerows(rows)

def classify_excel(input_xlsx, output_xlsx):
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]
    cat1_idx = headers.index('카테고리1')
    cat2_idx = headers.index('카테고리2')
    # 새 워크북 생성
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(['카테고리1', '카테고리2', '대분류'])
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat1 = row[cat1_idx]
        cat2 = row[cat2_idx]
        big = classify(cat2 or '', cat1 or '')
        out_ws.append([cat1, cat2, big])
    out_wb.save(output_xlsx)

if __name__ == '__main__':
    main()
    # 중복 없는 조합을 CP949로 저장
    save_unique_cp949(output_path, os.path.join(os.path.dirname(__file__), '카테고리_분류결과3_cp949.csv'))
    # classify_excel('category_cornerlogis_wholelist.xlsx', 'category_cornerlogis_wholelist_classified.xlsx') 